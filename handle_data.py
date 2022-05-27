import openpyxl, time, cv2
import glovar
import numpy as np
output_file_path = '.\Out'
rawPath = r'D:\Python\Project\PyTool_HandleRAW\700.raw'
imgSize = (2064, 2672)
from PIL import Image as im

def creat_report_excel():
    name_time = time.strftime("%Y%m%d%H%M%S", time.localtime())
    report_file_name = 'TestReport_' + name_time + '.xlsx'
    save_path_name = output_file_path + '\\' + report_file_name
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Test_Result'
    ws['A1'] = 'DUT_NO'
    ws['B1'] = 'Result'
    ws['C1'] = 'Fail_Item'
    ws['D1'] = 'Pin_Name'
    ws['E1'] = 'Fail_Value'
    ws['F1'] = 'Image_Status'
    wb.save(save_path_name)
    print('Finish report initial')
    return wb,

def transfer_16bit_to_8bit():
    # image_16bit = cv2.imread(image_path, cv2.IMREAD_UNCHANGED)
    image_16bit = np.fromfile(rawPath, dtype='>u2')  # >u2
    image_16bit = np.reshape(image_16bit[32:], imgSize)
    min_16bit = np.min(image_16bit)
    max_16bit = np.max(image_16bit)
    # image_8bit = np.array(np.rint((255.0 * (image_16bit - min_16bit)) / float(max_16bit - min_16bit)), dtype=np.uint8)
    # 或者下面一种写法
    image_8bit = np.array(np.rint(255 * ((image_16bit - min_16bit) / (max_16bit - min_16bit))), dtype=np.uint8)
    # print(image_16bit.dtype)
    # print('16bit dynamic range: %d - %d' % (min_16bit, max_16bit))
    # print(image_8bit.dtype)
    # print('8bit dynamic range: %d - %d' % (np.min(image_8bit), np.max(image_8bit)))

    # img = im.fromarray(image_8bit)
    # img.show()
    return image_8bit, image_16bit

# 全局变量
g_window_name = "img"  # 窗口名
g_window_wh = [800, 600]  # 窗口宽高

g_location_win = [0, 0]  # 相对于大图，窗口在图片中的位置
location_win = [0, 0]  # 鼠标左键点击时，暂存g_location_win
g_location_click, g_location_release = [0, 0], [0, 0]  # 相对于窗口，鼠标左键点击和释放的位置
once_execute = True
g_zoom, g_step = 1, 1  # 图片缩放比例和缩放系数
# g_image_original = cv2.imread(glovar.display_file)  # 原始图片，建议大于窗口宽高（800*600）
# g_image_zoom = g_image_original.copy()  # 缩放后的图片
# g_image_show = g_image_original[g_location_win[1]:g_location_win[1] + g_window_wh[1],
#                g_location_win[0]:g_location_win[0] + g_window_wh[0]]  # 实际显示的图片


# 矫正窗口在图片中的位置
# img_wh:图片的宽高, win_wh:窗口的宽高, win_xy:窗口在图片的位置
def check_location(img_wh, win_wh, win_xy):
    for i in range(2):
        if win_xy[i] < 0:
            win_xy[i] = 0
        elif win_xy[i] + win_wh[i] > img_wh[i] and img_wh[i] > win_wh[i]:
            win_xy[i] = img_wh[i] - win_wh[i]
        elif win_xy[i] + win_wh[i] > img_wh[i] and img_wh[i] < win_wh[i]:
            win_xy[i] = 0
    # print(img_wh, win_wh, win_xy)


# 计算缩放倍数
# flag：鼠标滚轮上移或下移的标识, step：缩放系数，滚轮每步缩放0.1, zoom：缩放倍数
def count_zoom(flag, step, zoom):
    if flag > 0:  # 滚轮上移
        zoom += step
        if zoom > 1 + step * 2000:  # 最多只能放大到3倍
            zoom = 1 + step * 2000
    else:  # 滚轮下移
        zoom -= step
        if zoom < step:  # 最多只能缩小到0.1倍
            zoom = step
    zoom = round(zoom, 4)  # 取2位有效数字
    return zoom


# OpenCV鼠标事件
def mouse(event, x, y, flags, param):
    global g_location_click, g_location_release, g_location_win, location_win, g_zoom, once_execute#, g_image_show, g_image_zoom
    # print(glovar.display_file)
    # print(str(glovar.display_file))
    # print(g_window_wh[1])
    # g_image_original = cv2.imread(str(glovar.display_file))  # 原始图片，建议大于窗口宽高（800*600）
    g_image_original, g_image_original_16bit = transfer_16bit_to_8bit()
    g_image_zoom = g_image_original.copy()  # 缩放后的图片
    g_image_show = g_image_original[g_location_win[1]:g_location_win[1] + g_window_wh[1],
                   g_location_win[0]:g_location_win[0] + g_window_wh[0]]  # 实际显示的图片
    if event == cv2.EVENT_LBUTTONDOWN:  # 左键点击
        g_location_click = [x, y]  # 左键点击时，鼠标相对于窗口的坐标
        # print('g_location_click:', g_location_click)
        location_win = [g_location_win[0], g_location_win[1]]  # 窗口相对于图片的坐标，不能写成location_win = g_location_win
    elif event == cv2.EVENT_MOUSEMOVE and (flags & cv2.EVENT_FLAG_LBUTTON):  # 按住左键拖曳
        g_location_release = [x, y]  # 左键拖曳时，鼠标相对于窗口的坐标
        h1, w1 = g_image_zoom.shape[0:2]  # 缩放图片的宽高
        w2, h2 = g_window_wh  # 窗口的宽高
        show_wh = [0, 0]  # 实际显示图片的宽高
        if w1 < w2 and h1 < h2:  # 图片的宽高小于窗口宽高，无法移动
            show_wh = [w1, h1]
            g_location_win = [0, 0]
        elif w1 >= w2 and h1 < h2:  # 图片的宽度大于窗口的宽度，可左右移动
            show_wh = [w2, h1]
            g_location_win[0] = location_win[0] + g_location_click[0] - g_location_release[0]
        elif w1 < w2 and h1 >= h2:  # 图片的高度大于窗口的高度，可上下移动
            show_wh = [w1, h2]
            g_location_win[1] = location_win[1] + g_location_click[1] - g_location_release[1]
        else:  # 图片的宽高大于窗口宽高，可左右上下移动
            show_wh = [w2, h2]
            g_location_win[0] = location_win[0] + g_location_click[0] - g_location_release[0]
            g_location_win[1] = location_win[1] + g_location_click[1] - g_location_release[1]
        check_location([w1, h1], [w2, h2], g_location_win)  # 矫正窗口在图片中的位置
        g_image_show = g_image_zoom[g_location_win[1]:g_location_win[1] + show_wh[1],
                       g_location_win[0]:g_location_win[0] + show_wh[0]]  # 实际显示的图片
        print(g_image_original_16bit[x, y])
        print(x, y)
    elif event == cv2.EVENT_MOUSEWHEEL:  # 滚轮
        z = g_zoom  # 缩放前的缩放倍数，用于计算缩放后窗口在图片中的位置
        g_zoom = count_zoom(flags, g_step, g_zoom)  # 计算缩放倍数
        w1, h1 = [int(g_image_original.shape[1] * g_zoom), int(g_image_original.shape[0] * g_zoom)]  # 缩放图片的宽高
        w2, h2 = g_window_wh  # 窗口的宽高
        print('1111111')
        g_image_zoom = cv2.resize(g_image_original, (w1, h1), interpolation=cv2.INTER_NEAREST)  # 图片缩放
        print('2222222')
        # # g_image_zoom = cv2.resize(g_image_original, (w2, h2), fx=w1, fy=h1, interpolation=cv2.INTER_LINEAR)
        # show_wh = [w1, h1]
        # cv2.resizeWindow(g_window_name, w2, h2)
        show_wh = [0, 0]  # 实际显示图片的宽高
        if w1 < w2 and h1 < h2:  # 缩放后，图片宽高小于窗口宽高
            show_wh = [w1, h1]
            cv2.resizeWindow(g_window_name, w1, h1)
        elif w1 >= w2 and h1 < h2:  # 缩放后，图片高度小于窗口高度
            show_wh = [w2, h1]
            cv2.resizeWindow(g_window_name, w2, h1)
        elif w1 < w2 and h1 >= h2:  # 缩放后，图片宽度小于窗口宽度
            show_wh = [w1, h2]
            cv2.resizeWindow(g_window_name, w1, h2)
        else:  # 缩放后，图片宽高大于窗口宽高
            show_wh = [w2, h2]
            cv2.resizeWindow(g_window_name, w2, h2)
        g_location_win = [int((g_location_win[0] + x) * g_zoom / z - x),
                          int((g_location_win[1] + y) * g_zoom / z - y)]  # 缩放后，窗口在图片的位置
        check_location([w1, h1], [w2, h2], g_location_win)  # 矫正窗口在图片中的位置
        # print(g_location_win, show_wh)
        g_image_show = g_image_zoom[g_location_win[1]:g_location_win[1] + show_wh[1],
                       g_location_win[0]:g_location_win[0] + show_wh[0]]  # 实际的显示图片

    # if event == cv2.EVENT_MOUSEMOVE:
    #     print(g_image_original_16bit[x, y])
    #     print(x, y)
    #     message = '{}{}'.format(g_image_original_16bit[x, y], 'ttt')
    #     cv2.putText(g_image_original, message, (int(x), int(y)),
    #                 cv2.FONT_HERSHEY_SIMPLEX, 0.5, (139, 0, 0), 1)
    #     cv2.circle(g_image_original, (x, y), 1, (0, 0, 255), -1)
    cv2.imshow(g_window_name, g_image_show)

def mouse_operation():
    # print('glovar.x:', glovar.x)
    # print('glovar.display_file:', glovar.display_file)
    # 设置窗口
    cv2.namedWindow(g_window_name, cv2.WINDOW_NORMAL)
    # 设置窗口大小，只有当图片大于窗口时才能移动图片
    cv2.resizeWindow(g_window_name, g_window_wh[0], g_window_wh[1])
    cv2.moveWindow(g_window_name, 700, 100)  # 设置窗口在电脑屏幕中的位置
    # 鼠标事件的回调函数
    cv2.setMouseCallback(g_window_name, mouse)
    cv2.waitKey()  # 不可缺少，用于刷新图片，等待鼠标操作
    cv2.destroyAllWindows()


# 主函数
if __name__ == "__main__":
    # 设置窗口
    cv2.namedWindow(g_window_name, cv2.WINDOW_NORMAL)
    # 设置窗口大小，只有当图片大于窗口时才能移动图片
    cv2.resizeWindow(g_window_name, g_window_wh[0], g_window_wh[1])
    cv2.moveWindow(g_window_name, 700, 100)  # 设置窗口在电脑屏幕中的位置
    # 鼠标事件的回调函数
    cv2.setMouseCallback(g_window_name, mouse)
    cv2.waitKey()  # 不可缺少，用于刷新图片，等待鼠标操作
    cv2.destroyAllWindows()
